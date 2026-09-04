"""Alertas internos de certificados enviados pela aplicação WhatsContábil."""

import os
import re
from math import ceil
from datetime import datetime
from pathlib import Path
from time import sleep

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from xml.sax.saxutils import escape

from integracoes.whatscontabil import (
    ErroWhatsContabil,
    enviar_midia,
    enviar_mensagem_texto,
    enviar_template,
    normalizar_telefone_brasil,
)
from integracoes.google_drive import (
    ErroRelatorioDrive,
    enviar_relatorio_drive,
    obter_link_relatorio,
)
from registro_alertas import (
    alerta_ja_enviado,
    alerta_ja_possui_algum_envio,
    registrar_alerta_enviado,
    registrar_evento_envio,
)


DIAS_MINIMOS_ALERTA = 1
DIAS_MAXIMOS_ALERTA = 30
LIMITE_MENSAGEM = 2000
LIMITE_VARIAVEL_TEMPLATE = 900
INTERVALO_ENTRE_TEMPLATES = 8.0
SEPARADOR_RESUMO = " | "
TEMPLATE_CLIENTE = "aviso_vencimento_certificado"
TEMPLATE_EQUIPE_RESUMO = "resumo_pendencias_certificados"
TEMPLATE_RESPONSAVEL_RESUMO = "resumo_renovacoes_responsavel"
TEMPLATE_EQUIPE_DOCUMENTO = "relatorio_pendencias_certificados"
TEMPLATE_RESPONSAVEL_DOCUMENTO = "relatorio_renovacoes_responsavel"
TEMPLATE_ABERTURA_RELATORIO = "aviso_relatorio_certificados"


def _chave_cliente(alerta):
    cnpj = "".join(
        caractere
        for caractere in str(alerta.get("cnpj") or "")
        if caractere.isdigit()
    )
    empresa = re.sub(
        r"\s+",
        " ",
        str(alerta.get("empresa") or alerta.get("arquivo") or "").strip(),
    ).casefold()
    return f"cnpj:{cnpj}" if len(cnpj) == 14 else f"empresa:{empresa}"


def preparar_alertas_internos(resultados):
    """Prepara avisos de clientes e pendências cadastrais da equipe."""
    candidatos = []
    contatos_por_cliente = {}
    for resultado in resultados:
        dias = resultado.get("dias")
        if not isinstance(dias, int) or dias < 0:
            continue

        elegivel_cliente = DIAS_MINIMOS_ALERTA <= dias <= DIAS_MAXIMOS_ALERTA
        alerta = {
            "cnpj": resultado.get("cnpj"),
            "empresa": resultado.get("empresa"),
            "arquivo": resultado.get("arquivo"),
            "vencimento": resultado.get("vencimento"),
            "dias": dias,
            "email": resultado.get("email"),
            "dados_cliente": resultado.get("dados_cliente"),
            "elegivel_cliente": elegivel_cliente,
            "tipo_aviso": (
                f"aviso_{dias}"
                if dias in {30, 15}
                else "recuperacao"
                if elegivel_cliente
                else "pendencia_cadastro"
            ),
        }

        chave = _chave_cliente(alerta)
        contatos = contatos_por_cliente.setdefault(
            chave,
            {"possui_email": False, "possui_telefone": False},
        )
        contatos["possui_email"] = contatos["possui_email"] or bool(
            _separar_emails_validos(alerta.get("email"))
        )
        contatos["possui_telefone"] = contatos["possui_telefone"] or bool(
            _obter_telefones_validos(alerta)
        )
        candidatos.append((alerta, chave))

    alertas = []
    for alerta, chave in candidatos:
        contatos = contatos_por_cliente[chave]
        alerta["possui_email_consolidado"] = contatos["possui_email"]
        alerta["possui_telefone_consolidado"] = contatos["possui_telefone"]
        if (
            alerta["elegivel_cliente"]
            or not contatos["possui_email"]
            or not contatos["possui_telefone"]
        ):
            alertas.append(alerta)
    return alertas


def _separar_emails_validos(valor):
    padrao = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
    emails = []
    for parte in re.split(r"[;,]", str(valor or "")):
        email = parte.strip().casefold()
        if email and padrao.fullmatch(email) and email not in emails:
            emails.append(email)
    return emails


def _obter_telefones_validos(alerta):
    dados_cliente = alerta.get("dados_cliente") or {}
    ddd_padrao = dados_cliente.get("ddd")
    candidatos = [dados_cliente.get("telefone")]
    candidatos.extend(
        contato.get("telefone")
        for contato in dados_cliente.get("contatos", [])
    )
    telefones = []
    for candidato in candidatos:
        try:
            telefone = normalizar_telefone_brasil(candidato, ddd_padrao)
        except ErroWhatsContabil:
            continue
        if telefone not in telefones:
            telefones.append(telefone)
    return telefones


def classificar_alertas_por_contato(alertas):
    """Separa empresas com telefone das que exigem pesquisa pela equipe."""
    com_contato = []
    para_equipe = []
    pendencias_por_cliente = {}

    for alerta_original in alertas:
        alerta = dict(alerta_original)
        alerta["emails_validos"] = _separar_emails_validos(alerta.get("email"))
        alerta["telefones_validos"] = _obter_telefones_validos(alerta)
        alerta["dados_faltantes"] = []
        possui_telefone = alerta.get(
            "possui_telefone_consolidado",
            bool(alerta["telefones_validos"]),
        )
        possui_email = alerta.get(
            "possui_email_consolidado",
            bool(alerta["emails_validos"]),
        )
        if not possui_telefone:
            alerta["dados_faltantes"].append("telefone")
        if not possui_email:
            alerta["dados_faltantes"].append("e-mail")

        # O WhatsApp depende somente de telefone. O e-mail continua no
        # relatório como informação complementar, mas não bloqueia o envio.
        elegivel_cliente = alerta.get("elegivel_cliente")
        if elegivel_cliente is None:
            dias = alerta.get("dias")
            elegivel_cliente = (
                isinstance(dias, int)
                and DIAS_MINIMOS_ALERTA <= dias <= DIAS_MAXIMOS_ALERTA
            )

        if elegivel_cliente and alerta["telefones_validos"]:
            com_contato.append(alerta)

        # Os grupos podem se sobrepor: com telefone a empresa pode receber o
        # WhatsApp, mas qualquer dado ausente ainda deve ser corrigido.
        if alerta["dados_faltantes"]:
            chave = _chave_cliente(alerta)

            existente = pendencias_por_cliente.get(chave)
            if existente is None:
                pendencias_por_cliente[chave] = alerta
                para_equipe.append(alerta)
            else:
                for dado in alerta["dados_faltantes"]:
                    if dado not in existente["dados_faltantes"]:
                        existente["dados_faltantes"].append(dado)

    return com_contato, para_equipe


def _formatar_cnpj(cnpj):
    numeros = "".join(caractere for caractere in str(cnpj or "") if caractere.isdigit())
    if len(numeros) != 14:
        return str(cnpj or "não informado")
    return (
        f"{numeros[:2]}.{numeros[2:5]}.{numeros[5:8]}/"
        f"{numeros[8:12]}-{numeros[12:]}"
    )


def _formatar_vencimento(vencimento):
    if hasattr(vencimento, "strftime"):
        return vencimento.strftime("%d/%m/%Y")
    texto = str(vencimento or "").strip()
    for formato in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto, formato).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return texto or "nao informado"


def _formatar_telefone_exibicao(telefone):
    digitos = re.sub(r"\D", "", str(telefone or ""))
    if digitos.startswith("55") and len(digitos) in {12, 13}:
        digitos = digitos[2:]
    if len(digitos) == 11:
        return f"({digitos[:2]}) {digitos[2:7]}-{digitos[7:]}"
    if len(digitos) == 10:
        return f"({digitos[:2]}) {digitos[2:6]}-{digitos[6:]}"
    return str(telefone or "Nao localizado")


def _telefones_para_relatorio(alerta):
    telefones = alerta.get("telefones_validos") or _obter_telefones_validos(alerta)
    if not telefones:
        return "Nao localizado"
    return "; ".join(_formatar_telefone_exibicao(item) for item in telefones)


def _emails_para_relatorio(alerta):
    emails = alerta.get("emails_validos") or _separar_emails_validos(
        alerta.get("email")
    )
    return "; ".join(emails) if emails else "Nao localizado"


def criar_relatorio_excel(alertas, tipo, pasta_destino):
    """Cria um relatorio simples para anexar ao template da WhatsContabil."""
    if tipo not in {"responsavel", "equipe"}:
        raise ValueError("Tipo de relatorio da WhatsContabil invalido.")

    pasta_destino = Path(pasta_destino)
    pasta_destino.mkdir(parents=True, exist_ok=True)
    carimbo = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    caminho = pasta_destino / f"relatorio_{tipo}_{carimbo}.xlsx"

    planilha = Workbook()
    aba = planilha.active
    aba.title = "Renovacoes" if tipo == "responsavel" else "Pendencias"
    titulo = (
        "Certificados proximos do vencimento"
        if tipo == "responsavel"
        else "Pendencias de contato dos clientes"
    )

    cabecalhos = ["Empresa", "CNPJ", "Telefone(s)", "E-mail"]
    if tipo == "responsavel":
        cabecalhos.extend(["Vencimento", "Dias restantes"])
    else:
        cabecalhos.append("Dados faltantes")

    ultima_coluna = len(cabecalhos)
    aba.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_coluna)
    celula_titulo = aba.cell(1, 1, titulo)
    celula_titulo.font = Font(bold=True, color="FFFFFF", size=14)
    celula_titulo.fill = PatternFill("solid", fgColor="1F4E78")
    celula_titulo.alignment = Alignment(horizontal="center", vertical="center")
    aba.row_dimensions[1].height = 26

    aba.cell(2, 1, f"Gerado em {datetime.now().strftime('%d/%m/%Y as %H:%M')}")
    aba.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ultima_coluna)
    aba.cell(2, 1).font = Font(italic=True, color="666666")

    for coluna, cabecalho in enumerate(cabecalhos, start=1):
        celula = aba.cell(4, coluna, cabecalho)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="2F75B5")
        celula.alignment = Alignment(horizontal="center", vertical="center")

    for numero_linha, alerta in enumerate(alertas, start=5):
        valores = [
            alerta.get("empresa") or "Empresa nao informada",
            _formatar_cnpj(alerta.get("cnpj")),
            _telefones_para_relatorio(alerta),
            _emails_para_relatorio(alerta),
        ]
        if tipo == "responsavel":
            valores.extend([
                _formatar_vencimento(alerta.get("vencimento")),
                alerta.get("dias"),
            ])
        else:
            valores.append(
                " e ".join(alerta.get("dados_faltantes", [])) or "Cadastro"
            )

        for coluna, valor in enumerate(valores, start=1):
            celula = aba.cell(numero_linha, coluna, valor)
            celula.alignment = Alignment(vertical="top", wrap_text=True)
            if numero_linha % 2 == 0:
                celula.fill = PatternFill("solid", fgColor="D9EAF7")

        # O openpyxl nǜo calcula automaticamente a altura das linhas. Esta
        # estimativa evita que nomes e e-mails longos fiquem cortados quando
        # o relat��rio for aberto ou enviado como documento.
        caracteres_por_coluna = [34, 18, 21, 30, 18, 14]
        quantidade_linhas = max(
            1,
            *(
                ceil(len(str(valor or "")) / caracteres_por_coluna[indice])
                for indice, valor in enumerate(valores)
            ),
        )
        aba.row_dimensions[numero_linha].height = max(
            20,
            15 * quantidade_linhas,
        )

    aba.freeze_panes = "A5"
    aba.auto_filter.ref = f"A4:{chr(64 + ultima_coluna)}{max(4, aba.max_row)}"
    larguras = [38, 20, 24, 34, 20, 16]
    for coluna, largura in enumerate(larguras[:ultima_coluna], start=1):
        aba.column_dimensions[chr(64 + coluna)].width = largura
    aba.sheet_view.showGridLines = False
    aba.sheet_view.zoomScale = 85
    aba.page_setup.orientation = "landscape"
    aba.page_setup.fitToWidth = 1
    aba.page_setup.fitToHeight = 0
    aba.sheet_properties.pageSetUpPr.fitToPage = True

    planilha.save(caminho)
    return caminho


def criar_relatorio_pdf(alertas, tipo, pasta_destino):
    """Cria o PDF anexado aos templates oficiais da WhatsContabil."""
    if tipo not in {"responsavel", "equipe"}:
        raise ValueError("Tipo de relatorio da WhatsContabil invalido.")

    pasta_destino = Path(pasta_destino)
    pasta_destino.mkdir(parents=True, exist_ok=True)
    carimbo = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    caminho = pasta_destino / f"relatorio_{tipo}_{carimbo}.pdf"

    titulo = (
        "Certificados próximos do vencimento"
        if tipo == "responsavel"
        else "Pendências de contato dos clientes"
    )
    cabecalhos = ["Empresa", "CNPJ", "Telefone(s)", "E-mail"]
    if tipo == "responsavel":
        cabecalhos.extend(["Vencimento", "Dias restantes"])
        larguras = [54, 31, 34, 58, 29, 22]
    else:
        cabecalhos.append("Dados faltantes")
        larguras = [59, 32, 37, 65, 38]

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloRelatorio",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=4 * mm,
    )
    estilo_resumo = ParagraphStyle(
        "ResumoRelatorio",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#555555"),
        spaceAfter=4 * mm,
    )
    estilo_cabecalho = ParagraphStyle(
        "CabecalhoTabela",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.white,
    )
    estilo_celula = ParagraphStyle(
        "CelulaTabela",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=6.8 if tipo == "responsavel" else 7.2,
        leading=8 if tipo == "responsavel" else 9,
        alignment=TA_LEFT,
        wordWrap="CJK",
    )

    tabela_dados = [
        [Paragraph(escape(cabecalho), estilo_cabecalho) for cabecalho in cabecalhos]
    ]
    for alerta in alertas:
        valores = [
            alerta.get("empresa") or "Empresa não informada",
            _formatar_cnpj(alerta.get("cnpj")),
            _telefones_para_relatorio(alerta),
            _emails_para_relatorio(alerta),
        ]
        if tipo == "responsavel":
            valores.extend([
                _formatar_vencimento(alerta.get("vencimento")),
                str(
                    alerta.get("dias")
                    if alerta.get("dias") is not None
                    else "Não informado"
                ),
            ])
        else:
            valores.append(
                " e ".join(alerta.get("dados_faltantes", [])) or "Cadastro"
            )
        tabela_dados.append(
            [Paragraph(escape(str(valor)), estilo_celula) for valor in valores]
        )

    documento = SimpleDocTemplate(
        str(caminho),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=14 * mm,
        title=titulo,
        author="Monitor de Certificados Digitais",
    )
    tabela = Table(
        tabela_dados,
        colWidths=[largura * mm for largura in larguras],
        repeatRows=1,
        hAlign="CENTER",
    )
    comandos_tabela = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F75B5")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3 if tipo == "responsavel" else 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3 if tipo == "responsavel" else 4),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#1F4E78")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#D7E2EA")),
    ]
    for linha in range(2, len(tabela_dados), 2):
        comandos_tabela.append(
            ("BACKGROUND", (0, linha), (-1, linha), colors.HexColor("#EAF3F8"))
        )
    tabela.setStyle(TableStyle(comandos_tabela))

    def desenhar_rodape(canvas, _documento):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D7E2EA"))
        canvas.line(10 * mm, 10 * mm, landscape(A4)[0] - 10 * mm, 10 * mm)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(10 * mm, 6 * mm, "Monitor de Certificados Digitais")
        canvas.drawRightString(
            landscape(A4)[0] - 10 * mm,
            6 * mm,
            f"Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    historia = [
        Paragraph(escape(titulo), estilo_titulo),
        Paragraph(
            escape(
                f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
                f" | Total de registros: {len(alertas)}"
            ),
            estilo_resumo,
        ),
        Spacer(1, 1 * mm),
        tabela,
    ]
    documento.build(
        historia,
        onFirstPage=desenhar_rodape,
        onLaterPages=desenhar_rodape,
    )
    return caminho


def _variaveis_template_vencimento(alerta):
    """Monta as cinco variaveis do template Utility aprovado."""
    return [
        "Cliente",
        str(alerta.get("empresa") or "Empresa nao informada"),
        _formatar_cnpj(alerta.get("cnpj")),
        _formatar_vencimento(alerta.get("vencimento")),
        str(alerta.get("dias") if alerta.get("dias") is not None else "nao informado"),
    ]


def _resumir_pendencias_equipe(alertas):
    """Resume todas as pendências em uma variável curta do template."""
    partes = []
    for indice, alerta in enumerate(alertas):
        empresa = str(alerta.get("empresa") or "Empresa nao informada").strip()
        faltantes = " e ".join(alerta.get("dados_faltantes", [])) or "cadastro"
        item = f"{indice + 1}) {empresa} - sem {faltantes}"
        candidato = SEPARADOR_RESUMO.join([*partes, item])
        if len(candidato) > LIMITE_VARIAVEL_TEMPLATE:
            restantes = len(alertas) - indice
            complemento = f"{SEPARADOR_RESUMO}... e mais {restantes} no painel"
            resumo = SEPARADOR_RESUMO.join(partes)
            limite = LIMITE_VARIAVEL_TEMPLATE - len(complemento)
            return resumo[:limite].rstrip() + complemento
        partes.append(item)
    return SEPARADOR_RESUMO.join(partes)


def _variaveis_template_resumo_equipe(alertas, nome_destinatario):
    """Monta as três variáveis do template Utility consolidado da equipe."""
    return [
        nome_destinatario,
        str(len(alertas)),
        _resumir_pendencias_equipe(alertas),
    ]


def _resumir_renovacoes_responsavel(alertas):
    """Resume empresas e prazos em uma única variável do template."""
    partes = []
    for indice, alerta in enumerate(alertas):
        empresa = str(alerta.get("empresa") or "Empresa nao informada").strip()
        dias = alerta.get("dias")
        prazo = str(dias) if isinstance(dias, int) else "nao informado"
        item = f"{indice + 1}) {empresa} - {prazo} dias"
        candidato = SEPARADOR_RESUMO.join([*partes, item])
        if len(candidato) > LIMITE_VARIAVEL_TEMPLATE:
            restantes = len(alertas) - indice
            complemento = f"{SEPARADOR_RESUMO}... e mais {restantes} no painel"
            resumo = SEPARADOR_RESUMO.join(partes)
            limite = LIMITE_VARIAVEL_TEMPLATE - len(complemento)
            return resumo[:limite].rstrip() + complemento
        partes.append(item)
    return SEPARADOR_RESUMO.join(partes)


def _variaveis_template_resumo_responsavel(alertas, nome_responsavel):
    """Monta as três variáveis do resumo Utility do responsável."""
    return [
        nome_responsavel,
        str(len(alertas)),
        _resumir_renovacoes_responsavel(alertas),
    ]


def _montar_bloco(alerta):
    vencimento = alerta.get("vencimento")
    if hasattr(vencimento, "strftime"):
        vencimento = vencimento.strftime("%d/%m/%Y")
    else:
        vencimento = str(vencimento or "não informado")

    return (
        f"Empresa: {alerta.get('empresa') or 'não informada'}\n"
        f"CNPJ: {_formatar_cnpj(alerta.get('cnpj'))}\n"
        f"Vencimento: {vencimento}\n"
        f"Prazo: {alerta.get('dias')} dias\n"
        "Ação: entrar em contato para iniciar a renovação."
    )


def _montar_mensagem_cliente(alerta):
    vencimento = alerta.get("vencimento")
    if hasattr(vencimento, "strftime"):
        vencimento = vencimento.strftime("%d/%m/%Y")
    else:
        vencimento = str(vencimento or "não informado")
    return (
        "*AVISO DE VENCIMENTO - CERTIFICADO DIGITAL*\n\n"
        "Olá,\n\n"
        f"O certificado digital e-CNPJ da empresa "
        f"{alerta.get('empresa') or 'não informada'}, CNPJ "
        f"{_formatar_cnpj(alerta.get('cnpj'))}, vencerá em {vencimento}.\n\n"
        f"Faltam {alerta.get('dias')} dias para o vencimento. "
        "Entre em contato com a Office Contábil para iniciar o processo "
        "de renovação.\n\n"
        "Atenciosamente,\nOffice Contábil"
    )


def _montar_bloco_responsavel(alerta):
    bloco = _montar_bloco(alerta)
    telefones = ", ".join(alerta.get("telefones_validos", []))
    emails = ", ".join(alerta.get("emails_validos", [])) or "não localizado"
    return f"{bloco}\nTelefone(s): {telefones}\nE-mail(s): {emails}"


def _montar_bloco_equipe(alerta):
    bloco = _montar_bloco(alerta)
    faltantes = " e ".join(alerta.get("dados_faltantes", []))
    return f"{bloco}\nPendência: localizar/atualizar {faltantes}."


def _compactar_variavel_template(texto):
    """Remove quebras de linha e limita o texto aceito pelo template."""
    compacto = re.sub(r"\s+", " ", str(texto or "")).strip()
    if len(compacto) <= LIMITE_VARIAVEL_TEMPLATE:
        return compacto
    return compacto[: LIMITE_VARIAVEL_TEMPLATE - 3].rstrip() + "..."


def _montar_relatorios_template(alertas, titulo, montar_bloco):
    """Agrupa blocos curtos sem quebrar as regras das variaveis da Meta."""
    prefixo = f"{titulo}: "
    pacotes = []
    texto_atual = prefixo
    itens_atuais = []

    for alerta in alertas:
        bloco = _compactar_variavel_template(montar_bloco(alerta))
        separador = " | " if itens_atuais else ""
        candidato = texto_atual + separador + bloco
        if len(candidato) > LIMITE_VARIAVEL_TEMPLATE and itens_atuais:
            pacotes.append((texto_atual, itens_atuais))
            texto_atual = prefixo + bloco
            itens_atuais = [alerta]
        else:
            texto_atual = candidato
            itens_atuais.append(alerta)

    if itens_atuais:
        pacotes.append((texto_atual, itens_atuais))
    return pacotes


def _montar_relatorios(alertas, titulo, montar_bloco):
    cabecalho = f"*{titulo}*\n\n"
    rodape = "\n\nMensagem automática do Monitor de Certificados."
    pacotes = []
    texto_atual = cabecalho
    alertas_atuais = []
    for alerta in alertas:
        bloco = montar_bloco(alerta)
        separador = "\n\n--------------------\n\n" if alertas_atuais else ""
        if len(texto_atual + separador + bloco + rodape) > LIMITE_MENSAGEM and alertas_atuais:
            pacotes.append((texto_atual + rodape, alertas_atuais))
            texto_atual = cabecalho + bloco
            alertas_atuais = [alerta]
        else:
            texto_atual += separador + bloco
            alertas_atuais.append(alerta)
    if alertas_atuais:
        pacotes.append((texto_atual + rodape, alertas_atuais))
    return pacotes


def montar_mensagens_internas(alertas):
    """Consolida alertas respeitando o limite de 2.000 caracteres."""
    cabecalho = "*AVISO INTERNO - CERTIFICADOS DIGITAIS*\n\n"
    rodape = "\n\nMensagem automática do Monitor de Certificados."
    pacotes = []
    texto_atual = cabecalho
    alertas_atuais = []

    for alerta in alertas:
        bloco = _montar_bloco(alerta)
        separador = "\n\n--------------------\n\n" if alertas_atuais else ""
        candidato = texto_atual + separador + bloco + rodape

        if len(candidato) > LIMITE_MENSAGEM and alertas_atuais:
            pacotes.append((texto_atual + rodape, alertas_atuais))
            texto_atual = cabecalho + bloco
            alertas_atuais = [alerta]
        else:
            texto_atual += separador + bloco
            alertas_atuais.append(alerta)

    if alertas_atuais:
        pacotes.append((texto_atual + rodape, alertas_atuais))

    return pacotes


def simular_alertas_internos(alertas, telefone):
    """Mostra os três fluxos sem chamar a API."""
    print("---------------------------")
    print("SIMULAÇÃO WHATSCONTÁBIL - NENHUMA MENSAGEM SERÁ ENVIADA")
    print(f"Número de teste: {telefone or 'não configurado'}")

    if not alertas:
        print("Nenhum aviso de cliente ou pendência da equipe foi encontrado.")
        return

    com_contato, para_equipe = classificar_alertas_por_contato(alertas)
    relatorios_equipe = _montar_relatorios(
        para_equipe,
        "PENDENCIAS DE CONTATO PARA A EQUIPE",
        _montar_bloco_equipe,
    )
    print(f"Mensagens individuais de cliente: {len(com_contato)}")
    print(f"Resumos para o responsável: {1 if com_contato else 0}")
    print(f"Relatorios para a equipe: {len(relatorios_equipe)}")
    return
    print(f"Empresas com telefone válido: {len(com_contato)}")
    print(f"Empresas sem telefone encaminhadas para a equipe: {len(para_equipe)}")

    for alerta in com_contato:
        print("---------------------------")
        print("TESTE - MENSAGEM DESTINADA AO CLIENTE")
        print(_montar_mensagem_cliente(alerta))

    for mensagem, _ in _montar_relatorios(
        com_contato,
        "TESTE - RELATÓRIO PARA O FUNCIONÁRIO RESPONSÁVEL",
        _montar_bloco_responsavel,
    ):
        print(f"---------------------------\n{mensagem}")

    for mensagem, _ in _montar_relatorios(
        para_equipe,
        "TESTE - PENDÊNCIAS DE CONTATO PARA A EQUIPE",
        _montar_bloco_equipe,
    ):
        print(f"---------------------------\n{mensagem}")


def enviar_alertas_internos(alertas, pasta_projeto, telefone, whatsapp_id):
    """Executa cliente, responsável e equipe usando o número de teste."""
    modo = os.getenv("MODO_WHATSCONTABIL", "desativado").strip().casefold()
    if modo != "teste":
        raise ErroWhatsContabil(
            "Envio bloqueado: a WhatsContabil deve estar no modo de teste."
        )

    telefone_teste = os.getenv("WHATSCONTABIL_NUMERO_TESTE", "").strip()
    if not telefone_teste:
        raise ErroWhatsContabil(
            "Envio bloqueado: o numero de teste nao esta configurado."
        )

    # O argumento telefone e mantido por compatibilidade, mas nunca define o
    # destinatario. Isso impede que um telefone de cliente seja usado por engano.
    destinatario = normalizar_telefone_brasil(telefone_teste)
    com_contato, para_equipe = classificar_alertas_por_contato(alertas)
    ignorar_duplicidade_teste = os.getenv(
        "IGNORAR_DUPLICIDADE_WHATSCONTABIL_TESTE",
        "sim",
    ).strip().casefold() in {"1", "s", "sim", "true"}
    escopo_envio = os.getenv(
        "WHATSCONTABIL_ESCOPO_ENVIO_TESTE",
        "completo",
    ).strip().casefold()
    if escopo_envio not in {"nenhum", "relatorios", "clientes", "completo"}:
        raise ErroWhatsContabil("Escopo de envio de teste invalido.")
    resumo = {
        "enviados": 0,
        "duplicados": 0,
        "falhas": 0,
        "interrompidos": 0,
        "detalhes_falhas": [],
        "message_ids": [],
    }

    def registro_alerta(alerta, tipo):
        registro = dict(alerta)
        registro["email"] = f"whatsapp:{tipo}:{destinatario}"
        if alerta.get("tipo_aviso") == "recuperacao":
            registro["dias"] = 0
        return registro

    def filtrar_duplicados(itens, tipo):
        if ignorar_duplicidade_teste:
            return list(itens)
        pendentes = []
        for alerta in itens:
            registro = registro_alerta(alerta, tipo)
            ja_enviado = (
                alerta_ja_possui_algum_envio(registro)
                if alerta.get("tipo_aviso") == "recuperacao"
                else alerta_ja_enviado(registro)
            )
            if ja_enviado:
                resumo["duplicados"] += 1
                registrar_evento_envio(
                    registro,
                    tipo.replace("_teste", ""),
                    "duplicado",
                    "Envio ja registrado anteriormente",
                )
            else:
                pendentes.append(alerta)
        return pendentes

    clientes_envio = filtrar_duplicados(com_contato, "cliente_teste")
    responsavel_envio = filtrar_duplicados(com_contato, "responsavel_teste")
    equipe_envio = filtrar_duplicados(para_equipe, "equipe_teste")
    if escopo_envio == "nenhum":
        clientes_envio = []
        responsavel_envio = []
        equipe_envio = []
    elif escopo_envio == "relatorios":
        clientes_envio = []
    elif escopo_envio == "clientes":
        responsavel_envio = []
        equipe_envio = []

    nome_template = os.getenv("WHATSCONTABIL_TEMPLATE_TESTE", "").strip()
    if not nome_template:
        resumo["falhas"] = 1
        resumo["detalhes_falhas"].append({
            "empresa": "Monitor de Certificados",
            "email": f"whatsapp:{destinatario}",
            "erro": "Template oficial de certificados nao configurado",
        })
        print(
            "Envio oficial bloqueado: configure o nome do template "
            "aprovado em WHATSCONTABIL_TEMPLATE_TESTE."
        )
        return resumo

    nome_destinatario = os.getenv(
        "WHATSCONTABIL_NOME_DESTINATARIO_TESTE",
        "Equipe Office",
    ).strip() or "Equipe Office"
    nome_template_equipe = os.getenv(
        "WHATSCONTABIL_TEMPLATE_EQUIPE_TESTE",
        "",
    ).strip()
    nome_template_equipe_documento = os.getenv(
        "WHATSCONTABIL_TEMPLATE_EQUIPE_DOCUMENTO_TESTE",
        "",
    ).strip()
    nome_template_responsavel = os.getenv(
        "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_TESTE",
        "",
    ).strip()
    nome_template_responsavel_documento = os.getenv(
        "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_DOCUMENTO_TESTE",
        "",
    ).strip()
    nome_template_abertura_relatorio = os.getenv(
        "WHATSCONTABIL_TEMPLATE_ABERTURA_RELATORIO_TESTE",
        "",
    ).strip()
    nome_template_relatorio_link = os.getenv(
        "WHATSCONTABIL_TEMPLATE_RELATORIO_LINK_TESTE",
        "",
    ).strip()
    permitir_fallback_midia = os.getenv(
        "WHATSCONTABIL_PERMITIR_FALLBACK_MIDIA_TESTE",
        "nao",
    ).strip().casefold() in {"1", "s", "sim", "true"}
    nome_responsavel = os.getenv(
        "WHATSCONTABIL_NOME_RESPONSAVEL_TESTE",
        "Responsavel",
    ).strip() or "Responsavel"
    transmissoes = []
    pasta_relatorios = Path(pasta_projeto) / "relatorios" / "whatscontabil"

    def adicionar_relatorio_por_link(tipo, itens, nome_destino):
        """Gera, publica e prepara um relatorio; falhas nao cancelam os clientes."""
        try:
            caminho_relatorio = criar_relatorio_pdf(
                itens,
                tipo,
                pasta_relatorios,
            )
            arquivo_drive = enviar_relatorio_drive(
                caminho_relatorio,
                pasta_projeto,
            )
            link = obter_link_relatorio(arquivo_drive)
        except (ErroRelatorioDrive, OSError, ValueError) as erro:
            resumo["falhas"] += 1
            resumo["detalhes_falhas"].append({
                "empresa": f"Relatorio interno: {tipo}",
                "email": f"whatsapp:{destinatario}",
                "erro": str(erro),
            })
            print(f"Falha ao publicar o relatorio de {tipo} no Drive: {erro}")
            return

        descricao = (
            f"{len(itens)} certificados proximos do vencimento"
            if tipo == "responsavel"
            else f"{len(itens)} empresas com pendencias de contato"
        )
        transmissoes.append((
            tipo,
            nome_template_relatorio_link,
            [nome_destino, descricao, link],
            itens,
            None,
        ))

    if nome_template == TEMPLATE_CLIENTE:
        # No modo de teste, todos os avisos usam o mesmo telefone. Os
        # relatorios internos precisam ir primeiro para nao ficarem no fim de
        # um lote grande e serem recusados pela API por excesso de mensagens
        # seguidas ao mesmo destinatario.
        if nome_template_relatorio_link and responsavel_envio:
            adicionar_relatorio_por_link(
                "responsavel",
                responsavel_envio,
                nome_responsavel,
            )
        elif (
            nome_template_responsavel_documento
            == TEMPLATE_RESPONSAVEL_DOCUMENTO
            and responsavel_envio
        ):
            caminho_relatorio = criar_relatorio_pdf(
                responsavel_envio,
                "responsavel",
                pasta_relatorios,
            )
            transmissoes.append((
                "responsavel",
                nome_template_responsavel_documento,
                [nome_responsavel, str(len(responsavel_envio))],
                responsavel_envio,
                caminho_relatorio,
            ))
        elif (
            nome_template_responsavel == TEMPLATE_RESPONSAVEL_RESUMO
            and responsavel_envio
        ):
            transmissoes.append((
                "responsavel",
                nome_template_responsavel,
                _variaveis_template_resumo_responsavel(
                    responsavel_envio,
                    nome_responsavel,
                ),
                responsavel_envio,
                None,
            ))
        elif responsavel_envio:
            print(
                "Relatorio do responsavel nao enviado: configure o template "
                f"Utility com documento {TEMPLATE_RESPONSAVEL_DOCUMENTO} em "
                "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_DOCUMENTO_TESTE."
            )
        if nome_template_relatorio_link and equipe_envio:
            adicionar_relatorio_por_link(
                "equipe",
                equipe_envio,
                nome_destinatario,
            )
        elif (
            nome_template_equipe_documento == TEMPLATE_EQUIPE_DOCUMENTO
            and equipe_envio
        ):
            caminho_relatorio = criar_relatorio_pdf(
                equipe_envio,
                "equipe",
                pasta_relatorios,
            )
            transmissoes.append((
                "equipe",
                nome_template_equipe_documento,
                [nome_destinatario, str(len(equipe_envio))],
                equipe_envio,
                caminho_relatorio,
            ))
        elif nome_template_equipe == TEMPLATE_EQUIPE_RESUMO and equipe_envio:
            transmissoes.append((
                "equipe",
                nome_template_equipe,
                _variaveis_template_resumo_equipe(
                    equipe_envio,
                    nome_destinatario,
                ),
                equipe_envio,
                None,
            ))
        elif equipe_envio:
            print(
                "Relatorio da equipe nao enviado: configure o template "
                f"Utility com documento {TEMPLATE_EQUIPE_DOCUMENTO} em "
                "WHATSCONTABIL_TEMPLATE_EQUIPE_DOCUMENTO_TESTE."
            )

        for alerta in clientes_envio:
            transmissoes.append((
                "cliente",
                nome_template,
                _variaveis_template_vencimento(alerta),
                [alerta],
                None,
            ))
    else:
        for alerta in clientes_envio:
            transmissoes.append((
                "cliente",
                nome_template,
                [
                    "Cliente",
                    _compactar_variavel_template(_montar_mensagem_cliente(alerta)),
                ],
                [alerta],
                None,
            ))

        for mensagem, itens in _montar_relatorios_template(
            responsavel_envio,
            "RELATORIO PARA O FUNCIONARIO RESPONSAVEL",
            _montar_bloco_responsavel,
        ):
            transmissoes.append((
                "responsavel",
                nome_template,
                ["Funcionario responsavel", mensagem],
                itens,
                None,
            ))

        for mensagem, itens in _montar_relatorios_template(
            equipe_envio,
            "PENDENCIAS DE CONTATO PARA A EQUIPE",
            _montar_bloco_equipe,
        ):
            transmissoes.append((
                "equipe",
                nome_template,
                [nome_destinatario, mensagem],
                itens,
                None,
            ))

    print(
        f"Escopo das mensagens de teste: {escopo_envio}. "
        f"Templates preparados: {len(transmissoes)}"
    )
    falhas_consecutivas = 0
    limite_falhas = 3
    for numero, (
        tipo,
        template_transmissao,
        variaveis,
        itens,
        arquivo,
    ) in enumerate(
        transmissoes,
        start=1,
    ):
        if numero > 1:
            sleep(INTERVALO_ENTRE_TEMPLATES)
        try:
            usar_fallback_midia = (
                arquivo is not None
                and nome_template_abertura_relatorio
                == TEMPLATE_ABERTURA_RELATORIO
                and permitir_fallback_midia
            )
            if usar_fallback_midia:
                descricao_relatorio = (
                    "certificados proximos do vencimento"
                    if tipo == "responsavel"
                    else "pendencias nos dados de contato"
                )
                nome_abertura = (
                    nome_responsavel
                    if tipo == "responsavel"
                    else nome_destinatario
                )
                resultado = enviar_template(
                    pasta_projeto,
                    destinatario,
                    nome_template_abertura_relatorio,
                    whatsapp_id,
                    [nome_abertura, descricao_relatorio],
                )
                resposta = resultado.get("resposta") or {}
                resumo["message_ids"].extend(resposta.get("messageIds") or [])
                resumo["enviados"] += 1
                sleep(5.1)
                resultado = enviar_midia(
                    pasta_projeto,
                    destinatario,
                    f"Relatorio de {descricao_relatorio}.",
                    whatsapp_id,
                    arquivo,
                )
            else:
                resultado = enviar_template(
                    pasta_projeto,
                    destinatario,
                    template_transmissao,
                    whatsapp_id,
                    variaveis,
                    arquivo=arquivo,
                )
            resposta = resultado.get("resposta") or {}
            resumo["message_ids"].extend(resposta.get("messageIds") or [])
            resumo["enviados"] += 1
            falhas_consecutivas = 0
            for item in itens:
                registrar_alerta_enviado(registro_alerta(item, tipo + "_teste"))
            print(
                f"Template de {tipo} {numero}/{len(transmissoes)} aceito "
                f"pela WhatsContabil ({len(itens)} certificado(s)"
                f"{'; documento enviado como midia' if usar_fallback_midia else '; documento anexado' if arquivo else ''})."
            )
        except (ErroWhatsContabil, OSError) as erro:
            falhas_consecutivas += 1
            resumo["falhas"] += 1
            resumo["detalhes_falhas"].append({
                "empresa": ", ".join(
                    item.get("empresa") or "Empresa nao informada"
                    for item in itens
                ),
                "email": f"whatsapp:{destinatario}",
                "erro": str(erro),
            })
            for item in itens:
                registrar_evento_envio(
                    registro_alerta(item, tipo + "_teste"),
                    tipo,
                    "falhou",
                    str(erro),
                )
            print(f"Falha no template de {tipo}: {erro}")
            restantes = len(transmissoes) - numero
            if falhas_consecutivas >= limite_falhas and restantes:
                resumo["interrompidos"] += restantes
                for tipo_restante, _, _, itens_restantes, _ in transmissoes[numero:]:
                    for item in itens_restantes:
                        registrar_evento_envio(
                            registro_alerta(item, tipo_restante + "_teste"),
                            tipo_restante,
                            "interrompido",
                            "Lote interrompido apos falhas consecutivas",
                        )
                print(
                    "Lote interrompido por seguranca apos "
                    f"{falhas_consecutivas} falhas consecutivas. "
                    f"Mensagens nao tentadas: {restantes}."
                )
                break
            print("O lote continuara com o proximo template.")

    return resumo

    # A conexão usada no ambiente de teste é oficial (Cloud API da Meta).
    # Nesse tipo de conexão, texto livre não pode iniciar uma conversa. O
    # template aprovado é enviado como notificação resumida; os detalhes
    # permanecem disponíveis no Monitor de Certificados.
    if ignorar_duplicidade_teste:
        nome_template = os.getenv(
            "WHATSCONTABIL_TEMPLATE_TESTE",
            "",
        ).strip()
        if not nome_template:
            resumo["falhas"] = 1
            resumo["detalhes_falhas"].append(
                {
                    "empresa": "Resumo do monitor",
                    "email": f"whatsapp:{destinatario}",
                    "erro": "Template oficial de certificados nao configurado",
                }
            )
            print(
                "Envio oficial bloqueado: configure o nome do template "
                "aprovado em WHATSCONTABIL_TEMPLATE_TESTE."
            )
            return resumo
        nome_destinatario = os.getenv(
            "WHATSCONTABIL_NOME_DESTINATARIO_TESTE",
            "Equipe Office",
        ).strip() or "Equipe Office"
        assunto = (
            f"{len(alertas)} certificados digitais proximos do vencimento. "
            "Consulte os detalhes no Monitor de Certificados"
        )

        try:
            resultado = enviar_template(
                pasta_projeto,
                destinatario,
                nome_template,
                whatsapp_id,
                [nome_destinatario, assunto],
            )
            resposta = resultado.get("resposta") or {}
            print(
                "Notificacao oficial de teste aceita pela WhatsContabil: "
                f"{resposta.get('message') or 'HTTP 2xx'}."
            )
            resumo["enviados"] = 1
        except (ErroWhatsContabil, OSError) as erro:
            resumo["falhas"] = 1
            resumo["detalhes_falhas"].append(
                {
                    "empresa": "Resumo do monitor",
                    "email": f"whatsapp:{destinatario}",
                    "erro": str(erro),
                }
            )
            print(f"Falha no envio do template oficial de teste: {erro}")
        return resumo

    def separar_pendentes(itens, tipo):
        resultado = []
        for alerta in itens:
            registro = dict(alerta)
            registro["email"] = f"whatsapp:{tipo}:{destinatario}"
            eh_recuperacao = alerta.get("tipo_aviso") == "recuperacao"

            if not ignorar_duplicidade_teste:
                if eh_recuperacao and alerta_ja_possui_algum_envio(registro):
                    resumo["duplicados"] += 1
                    continue
                if not eh_recuperacao and alerta_ja_enviado(registro):
                    resumo["duplicados"] += 1
                    continue

            # O valor zero identifica a recuperação independentemente do
            # dia em que ela aconteceu. Os avisos normais mantêm 30 ou 15.
            if eh_recuperacao:
                registro["dias"] = 0
            resultado.append((alerta, registro))
        return resultado

    cliente_pendentes = separar_pendentes(com_contato, "cliente_teste")
    responsavel_pendentes = separar_pendentes(com_contato, "responsavel_teste")
    equipe_pendentes = separar_pendentes(para_equipe, "equipe_teste")
    transmissoes = []

    for alerta, registro in cliente_pendentes:
        transmissoes.append(
            (
                "cliente",
                _montar_mensagem_cliente(alerta),
                [(alerta, registro)],
            )
        )

    mapa_responsavel = {id(alerta): registro for alerta, registro in responsavel_pendentes}
    for mensagem, itens in _montar_relatorios(
        [alerta for alerta, _ in responsavel_pendentes],
        "RELATÓRIO PARA O FUNCIONÁRIO RESPONSÁVEL",
        _montar_bloco_responsavel,
    ):
        transmissoes.append(
            (
                "responsável",
                mensagem,
                [(alerta, mapa_responsavel[id(alerta)]) for alerta in itens],
            )
        )

    mapa_equipe = {id(alerta): registro for alerta, registro in equipe_pendentes}
    for mensagem, itens in _montar_relatorios(
        [alerta for alerta, _ in equipe_pendentes],
        "PENDÊNCIAS DE CONTATO PARA A EQUIPE",
        _montar_bloco_equipe,
    ):
        transmissoes.append(
            (
                "equipe",
                mensagem,
                [(alerta, mapa_equipe[id(alerta)]) for alerta in itens],
            )
        )

    print(f"Mensagens de teste preparadas: {len(transmissoes)}")
    for numero, (tipo, mensagem, itens_registro) in enumerate(transmissoes, start=1):
        if numero > 1:
            sleep(5.1)
        try:
            resultado = enviar_mensagem_texto(
                pasta_projeto,
                destinatario,
                mensagem,
                whatsapp_id,
            )
            resposta = resultado.get("resposta") or {}
            ids = resposta.get("messageIds") or []
            resumo["message_ids"].extend(ids)

            for _, registro in itens_registro:
                registrar_alerta_enviado(registro)
                resumo["enviados"] += 1
            print(
                f"Mensagem de {tipo} {numero}/{len(transmissoes)} aceita "
                f"pela WhatsContábil ({len(itens_registro)} certificado(s))."
            )
        except (ErroWhatsContabil, OSError) as erro:
            resumo["falhas"] += len(itens_registro)
            for alerta, _ in itens_registro:
                resumo["detalhes_falhas"].append(
                    {
                        "empresa": alerta.get("empresa"),
                        "email": f"whatsapp:{destinatario}",
                        "erro": str(erro),
                    }
                )
            print(f"Falha no envio de {tipo}: {erro}")

    return resumo
