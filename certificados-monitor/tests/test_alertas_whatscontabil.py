import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


PASTA_MOTOR = Path(__file__).resolve().parents[1] / "automation_engine"
if str(PASTA_MOTOR) not in sys.path:
    sys.path.insert(0, str(PASTA_MOTOR))

import alertas_whatscontabil


class AlertasWhatsContabilTestCase(unittest.TestCase):
    def setUp(self):
        self.alerta = {
            "cnpj": "12345678000199",
            "empresa": "EMPRESA TESTE",
            "arquivo": "empresa.pfx",
            "vencimento": "2026-09-15",
            "dias": 15,
            "email": "cliente@example.com",
            "tipo_aviso": "aviso_15",
            "dados_cliente": {"telefone": "85999999999"},
        }

    def test_pendencia_valida_fora_da_janela_vai_somente_para_equipe(self):
        resultados = [{
            **self.alerta,
            "dias": 90,
            "email": None,
            "dados_cliente": {},
        }]

        preparados = alertas_whatscontabil.preparar_alertas_internos(resultados)
        clientes, equipe = alertas_whatscontabil.classificar_alertas_por_contato(
            preparados
        )

        self.assertEqual(len(preparados), 1)
        self.assertEqual(clientes, [])
        self.assertEqual(len(equipe), 1)
        self.assertEqual(equipe[0]["tipo_aviso"], "pendencia_cadastro")

    def test_certificado_completo_fora_da_janela_nao_gera_aviso(self):
        resultados = [{**self.alerta, "dias": 90}]

        preparados = alertas_whatscontabil.preparar_alertas_internos(resultados)

        self.assertEqual(preparados, [])

    def test_relatorio_excel_do_responsavel_inclui_telefone(self):
        alerta = {
            **self.alerta,
            "telefones_validos": ["5585999999999"],
            "emails_validos": ["cliente@example.com"],
        }
        with tempfile.TemporaryDirectory() as pasta:
            caminho = alertas_whatscontabil.criar_relatorio_excel(
                [alerta],
                "responsavel",
                pasta,
            )
            planilha = load_workbook(caminho, read_only=True)
            aba = planilha["Renovacoes"]
            valores = [aba.cell(5, coluna).value for coluna in range(1, 7)]
            planilha.close()

        self.assertEqual(valores[0], "EMPRESA TESTE")
        self.assertEqual(valores[1], "12.345.678/0001-99")
        self.assertEqual(valores[2], "(85) 99999-9999")
        self.assertEqual(valores[3], "cliente@example.com")
        self.assertEqual(valores[4], "15/09/2026")
        self.assertEqual(valores[5], 15)

    def test_relatorio_pdf_do_responsavel_e_gerado(self):
        alerta = {
            **self.alerta,
            "telefones_validos": ["5585999999999"],
            "emails_validos": ["cliente@example.com"],
        }
        with tempfile.TemporaryDirectory() as pasta:
            caminho = alertas_whatscontabil.criar_relatorio_pdf(
                [alerta],
                "responsavel",
                pasta,
            )
            conteudo = caminho.read_bytes()

        self.assertEqual(caminho.suffix, ".pdf")
        self.assertTrue(conteudo.startswith(b"%PDF"))
        self.assertGreater(len(conteudo), 1000)

    def test_pendencia_de_certificado_vencido_nao_gera_aviso(self):
        resultados = [{
            **self.alerta,
            "dias": -1,
            "email": None,
            "dados_cliente": {},
        }]

        preparados = alertas_whatscontabil.preparar_alertas_internos(resultados)

        self.assertEqual(preparados, [])

    def test_pendencias_da_equipe_sao_consolidadas_por_cnpj(self):
        primeiro = {
            **self.alerta,
            "dias": 90,
            "email": None,
            "dados_cliente": {},
        }
        segundo = {
            **primeiro,
            "empresa": "OUTRO NOME DO MESMO CLIENTE",
        }

        preparados = alertas_whatscontabil.preparar_alertas_internos(
            [primeiro, segundo]
        )
        _, equipe = alertas_whatscontabil.classificar_alertas_por_contato(
            preparados
        )

        self.assertEqual(len(equipe), 1)

    def test_contato_completo_em_duplicata_elimina_pendencia_do_cnpj(self):
        incompleto = {
            **self.alerta,
            "dias": 90,
            "email": None,
            "dados_cliente": {},
        }
        completo = {
            **self.alerta,
            "dias": 90,
            "empresa": "OUTRO NOME DO MESMO CLIENTE",
        }

        preparados = alertas_whatscontabil.preparar_alertas_internos(
            [incompleto, completo]
        )
        clientes, equipe = alertas_whatscontabil.classificar_alertas_por_contato(
            preparados
        )

        self.assertEqual(preparados, [])
        self.assertEqual(clientes, [])
        self.assertEqual(equipe, [])

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(alertas_whatscontabil, "registrar_alerta_enviado")
    @patch.object(alertas_whatscontabil, "enviar_mensagem_texto")
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"message": "Mensagem enviada com sucesso!"}
    })
    @patch.object(alertas_whatscontabil, "alerta_ja_enviado", return_value=True)
    def test_agendamento_de_teste_ignora_duplicidade(
        self,
        _duplicado,
        enviar_template_oficial,
        enviar_texto,
        _registrar,
        _sleep,
    ):
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "558596490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "alerta_certificado_teste",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [self.alerta],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        self.assertEqual(enviar_template_oficial.call_count, 2)
        for chamada in enviar_template_oficial.call_args_list:
            self.assertEqual(chamada.args[1], "558596490370")
        self.assertEqual(
            enviar_template_oficial.call_args_list[0].args[4][0],
            "Cliente",
        )
        for chamada in enviar_template_oficial.call_args_list:
            mensagem = chamada.args[4][1]
            self.assertNotIn("\n", mensagem)
            self.assertLessEqual(
                len(mensagem),
                alertas_whatscontabil.LIMITE_VARIAVEL_TEMPLATE,
            )
        self.assertEqual(
            enviar_template_oficial.call_args_list[1].args[4][0],
            "Funcionario responsavel",
        )
        enviar_texto.assert_not_called()
        self.assertEqual(resumo["enviados"], 2)
        self.assertEqual(resumo["duplicados"], 0)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"message": "Mensagem enviada com sucesso!"}
    })
    def test_empresa_sem_contato_vai_somente_para_equipe(
        self,
        enviar_template_oficial,
        _sleep,
    ):
        alerta_sem_contato = {
            **self.alerta,
            "email": None,
            "dados_cliente": {},
        }
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "558596490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "alerta_certificado_teste",
                "WHATSCONTABIL_NOME_DESTINATARIO_TESTE": "Equipe Office",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [alerta_sem_contato],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        enviar_template_oficial.assert_called_once()
        chamada = enviar_template_oficial.call_args
        self.assertEqual(chamada.args[1], "558596490370")
        self.assertEqual(chamada.args[4][0], "Equipe Office")
        self.assertIn("PENDENCIAS DE CONTATO", chamada.args[4][1])
        self.assertEqual(resumo["enviados"], 1)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"message": "Mensagem enviada com sucesso!"}
    })
    def test_template_utility_recebe_as_cinco_variaveis_aprovadas(
        self,
        enviar_template_oficial,
        _sleep,
    ):
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "5585996490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "aviso_vencimento_certificado",
                "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_TESTE": "",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [self.alerta],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        enviar_template_oficial.assert_called_once()
        variaveis = enviar_template_oficial.call_args.args[4]
        self.assertEqual(
            variaveis,
            [
                "Cliente",
                "EMPRESA TESTE",
                "12.345.678/0001-99",
                "15/09/2026",
                "15",
            ],
        )
        self.assertEqual(resumo["enviados"], 1)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"message": "Mensagem enviada com sucesso!"}
    })
    def test_responsavel_recebe_um_unico_resumo_consolidado(
        self,
        enviar_template_oficial,
        _sleep,
    ):
        segundo_alerta = {
            **self.alerta,
            "cnpj": "98765432000110",
            "empresa": "SEGUNDA EMPRESA",
            "dias": 22,
        }
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "5585996490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "aviso_vencimento_certificado",
                "WHATSCONTABIL_TEMPLATE_EQUIPE_TESTE": "resumo_pendencias_certificados",
                "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_TESTE": "resumo_renovacoes_responsavel",
                "WHATSCONTABIL_NOME_RESPONSAVEL_TESTE": "Responsavel interno",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [self.alerta, segundo_alerta],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        self.assertEqual(enviar_template_oficial.call_count, 3)
        chamadas_responsavel = [
            chamada
            for chamada in enviar_template_oficial.call_args_list
            if chamada.args[2] == "resumo_renovacoes_responsavel"
        ]
        self.assertEqual(len(chamadas_responsavel), 1)
        variaveis = chamadas_responsavel[0].args[4]
        self.assertEqual(variaveis[0], "Responsavel interno")
        self.assertEqual(variaveis[1], "2")
        self.assertEqual(
            variaveis[2],
            "1) EMPRESA TESTE - 15 dias | 2) SEGUNDA EMPRESA - 22 dias",
        )
        self.assertEqual(resumo["enviados"], 3)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"message": "Mensagem enviada com sucesso!"}
    })
    def test_template_utility_da_equipe_recebe_um_resumo_com_tres_variaveis(
        self,
        enviar_template_oficial,
        _sleep,
    ):
        alerta_sem_contato = {
            **self.alerta,
            "email": None,
            "dados_cliente": {},
        }
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "5585996490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "aviso_vencimento_certificado",
                "WHATSCONTABIL_TEMPLATE_EQUIPE_TESTE": "resumo_pendencias_certificados",
                "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_TESTE": "",
                "WHATSCONTABIL_NOME_DESTINATARIO_TESTE": "Equipe Office",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [alerta_sem_contato],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        enviar_template_oficial.assert_called_once()
        chamada = enviar_template_oficial.call_args
        self.assertEqual(chamada.args[2], "resumo_pendencias_certificados")
        self.assertEqual(
            chamada.args[4],
            [
                "Equipe Office",
                "1",
                "1) EMPRESA TESTE - sem telefone e e-mail",
            ],
        )
        self.assertEqual(resumo["enviados"], 1)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(
        alertas_whatscontabil,
        "criar_relatorio_pdf",
        side_effect=[Path("responsavel.pdf"), Path("equipe.pdf")],
    )
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"messageIds": ["mensagem-1"]}
    })
    def test_resumos_usam_templates_com_documento_quando_configurados(
        self,
        enviar_template_oficial,
        criar_relatorio,
        _sleep,
    ):
        alerta_sem_contato = {
            **self.alerta,
            "cnpj": "98765432000110",
            "empresa": "EMPRESA SEM CONTATO",
            "email": None,
            "dados_cliente": {},
        }
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "5585996490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "aviso_vencimento_certificado",
                "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_DOCUMENTO_TESTE": (
                    "relatorio_renovacoes_responsavel"
                ),
                "WHATSCONTABIL_TEMPLATE_EQUIPE_DOCUMENTO_TESTE": (
                    "relatorio_pendencias_certificados"
                ),
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [self.alerta, alerta_sem_contato],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        self.assertEqual(criar_relatorio.call_count, 2)
        self.assertEqual(enviar_template_oficial.call_count, 3)
        chamadas = enviar_template_oficial.call_args_list
        self.assertIsNone(chamadas[0].kwargs["arquivo"])
        self.assertEqual(chamadas[1].kwargs["arquivo"], Path("responsavel.pdf"))
        self.assertEqual(chamadas[2].kwargs["arquivo"], Path("equipe.pdf"))
        self.assertEqual(chamadas[1].args[4][1], "1")
        self.assertEqual(chamadas[2].args[4][1], "1")
        self.assertEqual(resumo["enviados"], 3)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(
        alertas_whatscontabil,
        "criar_relatorio_pdf",
        return_value=Path("responsavel.pdf"),
    )
    @patch.object(alertas_whatscontabil, "enviar_midia", return_value={
        "resposta": {"messageIds": ["midia-1"]}
    })
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"messageIds": ["template-1"]}
    })
    def test_relatorio_pode_usar_template_de_abertura_e_midia(
        self,
        enviar_template_oficial,
        enviar_midia_oficial,
        criar_relatorio,
        sleep_mock,
    ):
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "5585996490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "aviso_vencimento_certificado",
                "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_DOCUMENTO_TESTE": (
                    "relatorio_renovacoes_responsavel"
                ),
                "WHATSCONTABIL_TEMPLATE_EQUIPE_DOCUMENTO_TESTE": "",
                "WHATSCONTABIL_TEMPLATE_ABERTURA_RELATORIO_TESTE": (
                    "aviso_relatorio_certificados"
                ),
                "WHATSCONTABIL_NOME_RESPONSAVEL_TESTE": "Responsavel interno",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [self.alerta],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        self.assertEqual(criar_relatorio.call_count, 1)
        self.assertEqual(enviar_template_oficial.call_count, 2)
        abertura = enviar_template_oficial.call_args_list[1]
        self.assertEqual(abertura.args[2], "aviso_relatorio_certificados")
        self.assertEqual(
            abertura.args[4],
            ["Responsavel interno", "certificados proximos do vencimento"],
        )
        enviar_midia_oficial.assert_called_once_with(
            PASTA_MOTOR,
            "5585996490370",
            "Relatorio de certificados proximos do vencimento.",
            2,
            Path("responsavel.pdf"),
        )
        self.assertEqual(resumo["enviados"], 3)
        self.assertEqual(sleep_mock.call_count, 2)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"message": "Mensagem enviada com sucesso!"}
    })
    def test_empresa_com_telefone_e_sem_email_gera_cliente_e_equipe(
        self,
        enviar_template_oficial,
        _sleep,
    ):
        alerta_sem_email = {**self.alerta, "email": None}
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "5585996490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "aviso_vencimento_certificado",
                "WHATSCONTABIL_TEMPLATE_EQUIPE_TESTE": "resumo_pendencias_certificados",
                "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_TESTE": "",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [alerta_sem_email],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        self.assertEqual(enviar_template_oficial.call_count, 2)
        self.assertEqual(
            [chamada.args[2] for chamada in enviar_template_oficial.call_args_list],
            ["aviso_vencimento_certificado", "resumo_pendencias_certificados"],
        )
        self.assertEqual(resumo["enviados"], 2)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(alertas_whatscontabil, "enviar_template", return_value={
        "resposta": {"message": "Mensagem enviada com sucesso!"}
    })
    def test_varias_pendencias_geram_uma_unica_mensagem_da_equipe(
        self,
        enviar_template_oficial,
        _sleep,
    ):
        primeira = {**self.alerta, "email": None, "dados_cliente": {}}
        segunda = {
            **primeira,
            "cnpj": "98765432000110",
            "empresa": "SEGUNDA EMPRESA",
        }
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "5585996490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "aviso_vencimento_certificado",
                "WHATSCONTABIL_TEMPLATE_EQUIPE_TESTE": "resumo_pendencias_certificados",
                "WHATSCONTABIL_TEMPLATE_RESPONSAVEL_TESTE": "",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [primeira, segunda],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        enviar_template_oficial.assert_called_once()
        variaveis = enviar_template_oficial.call_args.args[4]
        self.assertEqual(variaveis[1], "2")
        self.assertEqual(
            variaveis[2],
            "1) EMPRESA TESTE - sem telefone e e-mail | "
            "2) SEGUNDA EMPRESA - sem telefone e e-mail",
        )
        self.assertEqual(resumo["enviados"], 1)

    @patch.object(alertas_whatscontabil, "sleep")
    @patch.object(
        alertas_whatscontabil,
        "enviar_template",
        side_effect=alertas_whatscontabil.ErroWhatsContabil(
            "A API recusou os parametros da requisicao."
        ),
    )
    def test_erro_em_um_template_nao_cancela_as_demais_tentativas(
        self,
        enviar_template_oficial,
        _sleep,
    ):
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "558596490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "alerta_certificado_teste",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [self.alerta, {**self.alerta, "empresa": "EMPRESA DOIS"}],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        self.assertEqual(enviar_template_oficial.call_count, 3)
        self.assertEqual(resumo["enviados"], 0)
        self.assertEqual(resumo["falhas"], 3)

    @patch.object(alertas_whatscontabil, "enviar_mensagem_texto")
    @patch.object(alertas_whatscontabil, "alerta_ja_enviado", return_value=True)
    def test_fora_do_modo_teste_bloqueia_envio(self, _duplicado, enviar):
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "desativado",
                "WHATSCONTABIL_NUMERO_TESTE": "558596490370",
            },
            clear=False,
        ):
            with self.assertRaises(alertas_whatscontabil.ErroWhatsContabil):
                alertas_whatscontabil.enviar_alertas_internos(
                    [self.alerta],
                    PASTA_MOTOR,
                    "5585999999999",
                    2,
                )

        enviar.assert_not_called()

    @patch.object(alertas_whatscontabil, "enviar_mensagem_texto")
    @patch.object(alertas_whatscontabil, "enviar_template")
    def test_sem_template_aprovado_nao_usa_texto_livre(
        self,
        enviar_template_oficial,
        enviar_texto,
    ):
        with patch.dict(
            os.environ,
            {
                "MODO_WHATSCONTABIL": "teste",
                "WHATSCONTABIL_NUMERO_TESTE": "558596490370",
                "WHATSCONTABIL_TEMPLATE_TESTE": "",
            },
            clear=False,
        ):
            resumo = alertas_whatscontabil.enviar_alertas_internos(
                [self.alerta],
                PASTA_MOTOR,
                "5585999999999",
                2,
            )

        enviar_template_oficial.assert_not_called()
        enviar_texto.assert_not_called()
        self.assertEqual(resumo["falhas"], 1)


if __name__ == "__main__":
    unittest.main()
